"""
Reader Module - Excel File Import and Sheet Handling
====================================================

Handles importing Excel files with intelligent structure detection,
merged cell handling, and multi-sheet support.
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional, Union

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def get_sheets(file_path: Union[str, Path]) -> List[str]:
    """
    Get a list of all sheet names in an Excel file.

    Args:
        file_path: Path to the Excel file

    Returns:
        List of sheet names

    Raises:
        FileNotFoundError: If the file doesn't exist
        ValueError: If the file is not a valid Excel file

    Example:
        >>> sheets = get_sheets("data.xlsx")
        >>> print(sheets)
        ['Sheet1', 'Sheet2', 'Summary']
    """
    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    if file_path.suffix not in ['.xlsx', '.xls', '.xlsm']:
        raise ValueError(f"Not a valid Excel file: {file_path}")

    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheets = workbook.sheetnames
        workbook.close()
        logger.info(f"Found {len(sheets)} sheets in {file_path.name}")
        return sheets
    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        raise ValueError(f"Could not read Excel file: {e}")


def _unmerge_cells(file_path: Union[str, Path], sheet_name: Optional[str] = None) -> pd.DataFrame:
    """
    Internal function to handle merged cells by filling them with the merged value.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to read

    Returns:
        DataFrame with unmerged cells
    """
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active

    # Get all merged cell ranges
    merged_ranges = list(worksheet.merged_cells.ranges)

    # Unmerge and fill cells
    for merged_range in merged_ranges:
        # Get the value from the top-left cell
        min_col, min_row, max_col, max_row = merged_range.bounds
        value = worksheet.cell(min_row, min_col).value

        # Unmerge the cells
        worksheet.unmerge_cells(str(merged_range))

        # Fill all cells in the range with the value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                worksheet.cell(row, col).value = value

    # Convert to DataFrame
    data = worksheet.values
    rows = list(data)

    workbook.close()

    if not rows:
        return pd.DataFrame()

    # Try to use first row as header if it looks like headers
    if rows and all(isinstance(cell, str) for cell in rows[0] if cell is not None):
        df = pd.DataFrame(rows[1:], columns=rows[0])
    else:
        df = pd.DataFrame(rows)

    return df


def import_excel(
        file_path: Union[str, Path],
        sheet_name: Optional[Union[str, int]] = None,
        handle_merged: bool = True,
        skip_rows: Optional[int] = None,
        **kwargs
) -> pd.DataFrame:
    """
    Import an Excel file with intelligent structure detection and merged cell handling.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name or index of the sheet to read (None for first sheet)
        handle_merged: Whether to automatically handle merged cells
        skip_rows: Number of rows to skip from the top
        **kwargs: Additional arguments passed to pandas.read_excel

    Returns:
        DataFrame containing the Excel data

    Raises:
        FileNotFoundError: If the file doesn't exist
        ValueError: If the file or sheet is invalid

    Example:
        >>> df = import_excel("data.xlsx", sheet_name="Objects")
        >>> print(df.head())
    """
    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    logger.info(f"Importing Excel file: {file_path.name}")

    try:
        # If we need to handle merged cells, use openpyxl
        if handle_merged:
            logger.info("Handling merged cells...")
            df = _unmerge_cells(file_path, sheet_name)

            if skip_rows:
                df = df.iloc[skip_rows:].reset_index(drop=True)
        else:
            # Use pandas for simpler cases
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name if sheet_name is not None else 0,
                skiprows=skip_rows,
                **kwargs
            )

        logger.info(f"Successfully imported {len(df)} rows and {len(df.columns)} columns")
        return df

    except Exception as e:
        logger.error(f"Error importing Excel file: {e}")
        raise ValueError(f"Could not import Excel file: {e}")


def preview_data(
        file_path: Union[str, Path],
        sheet_name: Optional[Union[str, int]] = None,
        rows: int = 5
) -> pd.DataFrame:
    """
    Preview the first N rows of an Excel sheet without loading the entire file.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name or index of the sheet to preview
        rows: Number of rows to preview (default: 5)

    Returns:
        DataFrame with preview data

    Example:
        >>> preview = preview_data("large_file.xlsx", rows=10)
        >>> print(preview)
    """
    logger.info(f"Previewing {rows} rows from {Path(file_path).name}")

    try:
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name if sheet_name is not None else 0,
            nrows=rows
        )
        return df
    except Exception as e:
        logger.error(f"Error previewing data: {e}")
        raise ValueError(f"Could not preview data: {e}")


def get_file_info(file_path: Union[str, Path]) -> Dict[str, any]:
    """
    Get comprehensive information about an Excel file.

    Args:
        file_path: Path to the Excel file

    Returns:
        Dictionary containing file information

    Example:
        >>> info = get_file_info("data.xlsx")
        >>> print(f"File has {info['sheet_count']} sheets")
    """
    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    sheets = get_sheets(file_path)

    info = {
        "filename": file_path.name,
        "file_size_mb": file_path.stat().st_size / (1024 * 1024),
        "sheet_count": len(sheets),
        "sheet_names": sheets,
    }

    # Get row/column counts for each sheet
    sheet_info = {}
    for sheet in sheets:
        try:
            df = preview_data(file_path, sheet_name=sheet, rows=1)
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb[sheet]
            sheet_info[sheet] = {
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "columns": list(df.columns) if not df.empty else []
            }
            wb.close()
        except Exception as e:
            logger.warning(f"Could not get info for sheet '{sheet}': {e}")
            sheet_info[sheet] = {"error": str(e)}

    info["sheets_info"] = sheet_info

    return info